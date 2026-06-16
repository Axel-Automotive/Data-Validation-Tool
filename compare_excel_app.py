import re

import streamlit as st
import pandas as pd
import io
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

st.set_page_config(
    page_title="Data Validator",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@400;500;600;700;800&display=swap');

:root {
    --p1: #7c3aed; --p2: #6d28d9; --p3: #4c1d95;
    --g1: #06b6d4; --g2: #0891b2;
    --ink: #0f0f23; --sub: #6b7280;
    --surf: #ffffff; --bg: #f4f4f8;
    --border: #e5e7eb;
    --side-bg: #0f0f23; --side-text: #a1a1c7; --side-accent: #7c3aed;
}

html, body, [class*="css"] {
    font-family: 'Plus Jakarta Sans', system-ui, sans-serif;
}

#MainMenu, footer, header { visibility: hidden; }

/* ── Background ── */
.stApp { background: var(--bg); }
.block-container { padding-top: 2rem; padding-bottom: 4rem; max-width: 1160px; }

/* ── Sidebar ── */
[data-testid="stSidebar"] {
    background: var(--side-bg) !important;
    border-right: none !important;
}
[data-testid="stSidebar"] > div:first-child { padding-top: 0 !important; }

/* sidebar labels */
.sb-section {
    font-size: 0.65rem; font-weight: 700; letter-spacing: 1.2px;
    text-transform: uppercase; color: #4b5563;
    margin: 1.6rem 0 0.7rem; display: block;
}

.sb-file-card {
    background: #1a1a3e;
    border: 1px solid #2d2d5e;
    border-radius: 10px;
    padding: 0.65rem 0.85rem;
    margin-top: 0.4rem;
    font-size: 0.78rem;
    display: flex; align-items: center; gap: 8px;
}
.sb-file-card.ok  { border-color: #4ade80; color: #4ade80; }
.sb-file-card.wait{ color: #4b5563; }

.sb-stat {
    font-size: 0.75rem; color: var(--side-text);
    background: #1a1a3e; border-radius: 8px;
    padding: 0.6rem 0.85rem; margin-top: 0.35rem;
    line-height: 1.8;
}

/* sidebar file uploader */
[data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] {
    background: #1a1a3e !important;
    border: 1.5px dashed #3d3d7e !important;
    border-radius: 10px !important;
}
[data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] p,
[data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] small {
    color: #6b6b9e !important;
}
[data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] svg { stroke: #6b6b9e !important; }

/* sidebar selectbox */
[data-testid="stSidebar"] [data-baseweb="select"] > div {
    background: #1a1a3e !important;
    border-color: #2d2d5e !important;
    color: #c4c4e8 !important;
    border-radius: 8px !important;
}
[data-testid="stSidebar"] label, [data-testid="stSidebar"] .stCaption p {
    color: var(--side-text) !important;
}

/* ── Page header ── */
.page-header {
    display: flex; align-items: flex-start;
    gap: 1.25rem; margin-bottom: 2rem;
}
.page-icon {
    width: 56px; height: 56px; border-radius: 16px; flex-shrink: 0;
    background: linear-gradient(135deg, var(--p1), var(--g1));
    display: flex; align-items: center; justify-content: center;
    font-size: 1.6rem; box-shadow: 0 8px 24px rgba(124,58,237,.35);
}
.page-title { flex: 1; }
.page-title h1 {
    font-size: 1.9rem; font-weight: 800; color: var(--ink);
    margin: 0; letter-spacing: -0.5px; line-height: 1.1;
}
.page-title p { font-size: 0.9rem; color: var(--sub); margin: 0.35rem 0 0; }
.header-divider {
    border: none; height: 1px;
    background: linear-gradient(90deg, #e5e7eb 0%, transparent 80%);
    margin: 0 0 0.5rem;
}

/* ── Metric cards ── */
.mrow { display: flex; gap: 0.8rem; flex-wrap: wrap; margin: 0.5rem 0 1.5rem; }
.mcard {
    flex: 1; min-width: 140px;
    background: var(--surf);
    border-radius: 16px;
    padding: 1.1rem 1.2rem 1rem;
    position: relative; overflow: hidden;
    box-shadow: 0 1px 3px rgba(0,0,0,.05), 0 4px 12px rgba(0,0,0,.04);
    transition: transform .18s, box-shadow .18s;
}
.mcard:hover { transform: translateY(-3px); box-shadow: 0 12px 28px rgba(0,0,0,.08); }
.mcard-accent {
    position: absolute; top: 0; left: 0; right: 0; height: 3px;
    background: var(--mc, #e5e7eb); border-radius: 16px 16px 0 0;
}
.mcard-icon {
    font-size: 1.25rem; margin-bottom: 0.6rem; display: block;
}
.mcard-label {
    font-size: 0.68rem; font-weight: 700; letter-spacing: 0.6px;
    text-transform: uppercase; color: var(--sub); margin-bottom: 0.3rem;
}
.mcard-value {
    font-size: 2rem; font-weight: 800; color: var(--ink);
    letter-spacing: -1.5px; line-height: 1;
}
.mcard-sub {
    font-size: 0.7rem; font-weight: 600; margin-top: 0.4rem;
    color: var(--mcs, var(--sub));
}

/* ── Tabs ── */
.stTabs [data-baseweb="tab-list"] {
    background: var(--surf); border: 1px solid var(--border);
    border-radius: 14px; padding: 5px; gap: 4px;
    box-shadow: 0 1px 3px rgba(0,0,0,.04);
}
.stTabs [data-baseweb="tab"] {
    border-radius: 10px; padding: 0.6rem 1.35rem;
    font-size: 0.875rem; font-weight: 600; color: var(--sub);
    transition: all .15s; background: transparent;
}
.stTabs [data-baseweb="tab"]:hover { color: var(--ink); background: var(--bg); }
.stTabs [aria-selected="true"] {
    background: linear-gradient(135deg, var(--p1), var(--p2)) !important;
    color: #fff !important;
    box-shadow: 0 4px 12px rgba(124,58,237,.4) !important;
}
.stTabs [data-baseweb="tab-highlight"],
.stTabs [data-baseweb="tab-border"] { display: none !important; }
.stTabs [data-baseweb="tab-panel"] { padding-top: 1.75rem; }

/* ── Section label ── */
.sec-label {
    font-size: 0.65rem; font-weight: 800; letter-spacing: 1px;
    text-transform: uppercase; color: var(--p1);
    display: flex; align-items: center; gap: 6px; margin: 0 0 0.9rem;
}
.sec-label::after {
    content: ""; flex: 1; height: 1px; background: var(--border);
}

/* ── Cards ── */
.card {
    background: var(--surf); border: 1px solid var(--border);
    border-radius: 16px; padding: 1.4rem 1.5rem;
    box-shadow: 0 1px 3px rgba(0,0,0,.04); margin-bottom: 1rem;
}

/* ── Buttons ── */
[data-testid="stButton"] button[kind="primary"] {
    background: linear-gradient(135deg, var(--p1), var(--p2)) !important;
    border: none !important; border-radius: 10px !important;
    font-weight: 700 !important; font-size: 0.875rem !important;
    padding: 0.65rem 2rem !important; color: #fff !important;
    box-shadow: 0 4px 16px rgba(124,58,237,.35) !important;
    letter-spacing: 0.2px; transition: all .15s !important;
}
[data-testid="stButton"] button[kind="primary"]:hover {
    box-shadow: 0 6px 24px rgba(124,58,237,.5) !important;
    transform: translateY(-1px);
}
[data-testid="stDownloadButton"] button {
    background: var(--surf) !important; color: var(--p1) !important;
    border: 1.5px solid #ddd6fe !important; border-radius: 10px !important;
    font-weight: 700 !important; padding: 0.6rem 1.5rem !important;
    transition: all .15s !important; font-size: 0.875rem !important;
}
[data-testid="stDownloadButton"] button:hover {
    background: #f5f3ff !important; border-color: var(--p1) !important;
    transform: translateY(-1px);
}

/* ── Inputs ── */
[data-testid="stTextInput"] input {
    border-radius: 10px !important; border-color: var(--border) !important;
    font-size: 0.875rem !important;
}
[data-testid="stTextInput"] input:focus { border-color: var(--p1) !important; box-shadow: 0 0 0 3px rgba(124,58,237,.1) !important; }
[data-baseweb="select"] > div {
    border-radius: 10px !important; border-color: var(--border) !important;
}
[data-testid="stMultiSelect"] div[data-baseweb="select"] > div {
    border-radius: 10px !important; border-color: var(--border) !important;
}

/* ── Expanders ── */
[data-testid="stExpander"] {
    border: 1px solid var(--border) !important; border-radius: 14px !important;
    background: var(--surf) !important; overflow: hidden;
    box-shadow: 0 1px 3px rgba(0,0,0,.04) !important;
}
[data-testid="stExpander"] summary {
    font-weight: 700 !important; font-size: 0.85rem !important;
    color: var(--ink) !important; padding: 0.9rem 1.1rem !important;
}
[data-testid="stExpander"] summary:hover { color: var(--p1) !important; }

/* ── DataFrame ── */
[data-testid="stDataFrame"] { border: 1px solid var(--border); border-radius: 12px; overflow: hidden; }

/* ── Alerts ── */
[data-testid="stAlert"] { border-radius: 12px !important; border-left-width: 4px !important; font-size: 0.875rem !important; }

/* ── File uploader (main) ── */
[data-testid="stFileUploaderDropzone"] {
    background: #fafafe !important; border: 1.5px dashed #c4b5fd !important;
    border-radius: 12px !important;
}

/* ── Feature cards (landing) ── */
.feat-grid { display: grid; grid-template-columns: repeat(3,1fr); gap: 1rem; margin-top: 1.5rem; }
.feat-card {
    background: var(--surf); border: 1px solid var(--border);
    border-radius: 20px; padding: 1.75rem 1.5rem;
    transition: transform .2s, box-shadow .2s;
    box-shadow: 0 1px 3px rgba(0,0,0,.04);
}
.feat-card:hover { transform: translateY(-4px); box-shadow: 0 16px 40px rgba(0,0,0,.1); }
.feat-icon {
    width: 52px; height: 52px; border-radius: 14px;
    display: flex; align-items: center; justify-content: center;
    font-size: 1.5rem; margin-bottom: 1.1rem;
}
.feat-card h3 { font-size: 1rem; font-weight: 800; color: var(--ink); margin: 0 0 0.5rem; }
.feat-card p  { font-size: 0.84rem; color: var(--sub); line-height: 1.7; margin: 0; }
.feat-badge {
    display: inline-block; margin-bottom: 0.9rem;
    font-size: 0.65rem; font-weight: 800; letter-spacing: 0.8px;
    text-transform: uppercase; padding: 0.22rem 0.65rem;
    border-radius: 30px; background: #f5f3ff; color: var(--p1);
}
</style>
""", unsafe_allow_html=True)


# ── Helpers ───────────────────────────────────────────────────────────────────
def sanitize_sheet_name(name, suffix="", max_len=31):
    name = re.sub(r'[/\\?*\[\]:]', '_', name)
    return name[:max_len - len(suffix)] + suffix


def mcard(label, value, sub="", color="#9ca3af", icon=""):
    sub_html = f'<div class="mcard-sub" style="--mcs:{color};">{sub}</div>' if sub else ""
    icon_html = f'<span class="mcard-icon">{icon}</span>' if icon else ""
    return (
        f'<div class="mcard">'
        f'<div class="mcard-accent" style="--mc:{color};"></div>'
        f'{icon_html}'
        f'<div class="mcard-label">{label}</div>'
        f'<div class="mcard-value">{value}</div>'
        f'{sub_html}'
        f'</div>'
    )


def metrics_row(*cards):
    st.markdown('<div class="mrow">' + "".join(cards) + '</div>', unsafe_allow_html=True)


def sec(title):
    st.markdown(f'<div class="sec-label">{title}</div>', unsafe_allow_html=True)


# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    try:
        st.image("vector.png", use_container_width=True)
    except Exception:
        pass

    st.markdown('<span class="sb-section">Source Files</span>', unsafe_allow_html=True)

    file_a = st.file_uploader("File A", type=["xlsx", "xls"], key="up_a",
                               label_visibility="collapsed")
    if file_a:
        st.markdown(f'<div class="sb-file-card ok">✓&nbsp; {file_a.name}</div>', unsafe_allow_html=True)
    else:
        st.markdown('<div class="sb-file-card wait">📄&nbsp; File A — not uploaded</div>', unsafe_allow_html=True)

    st.markdown("<div style='height:0.5rem'></div>", unsafe_allow_html=True)

    file_b = st.file_uploader("File B", type=["xlsx", "xls"], key="up_b",
                               label_visibility="collapsed")
    if file_b:
        st.markdown(f'<div class="sb-file-card ok">✓&nbsp; {file_b.name}</div>', unsafe_allow_html=True)
    else:
        st.markdown('<div class="sb-file-card wait">📄&nbsp; File B — not uploaded</div>', unsafe_allow_html=True)

    sheet_a = sheet_b = None
    df_a = df_b = None

    if file_a and file_b:
        bytes_a = file_a.getvalue()
        bytes_b = file_b.getvalue()
        excel_a = pd.ExcelFile(io.BytesIO(bytes_a))
        excel_b = pd.ExcelFile(io.BytesIO(bytes_b))

        st.markdown('<span class="sb-section">Sheet Selection</span>', unsafe_allow_html=True)
        sheet_a = st.selectbox(f"↳ {file_a.name}", excel_a.sheet_names, key="sheet_a_select")
        sheet_b = st.selectbox(f"↳ {file_b.name}", excel_b.sheet_names, key="sheet_b_select")

        df_a = excel_a.parse(sheet_name=sheet_a)
        df_b = excel_b.parse(sheet_name=sheet_b)

        st.markdown('<span class="sb-section">Dataset Info</span>', unsafe_allow_html=True)
        st.markdown(
            f'<div class="sb-stat">'
            f'<strong style="color:#a78bfa">A</strong> &nbsp;{len(df_a):,} rows · {len(df_a.columns)} cols<br>'
            f'<strong style="color:#67e8f9">B</strong> &nbsp;{len(df_b):,} rows · {len(df_b.columns)} cols'
            f'</div>',
            unsafe_allow_html=True,
        )


# ── Page Header ───────────────────────────────────────────────────────────────
st.markdown(
    '<div class="page-header">'
    '<div class="page-icon">⚡</div>'
    '<div class="page-title">'
    '<h1>Data Validator</h1>'
    '<p>Compare and reconcile Excel spreadsheets with precision</p>'
    '</div>'
    '</div>'
    '<hr class="header-divider">',
    unsafe_allow_html=True,
)


# ── Landing ───────────────────────────────────────────────────────────────────
if not file_a or not file_b:
    st.markdown(
        "<p style='font-size:0.9rem;color:#6b7280;margin-bottom:0.5rem;'>"
        "Upload two Excel files in the sidebar to unlock all comparison modes.</p>",
        unsafe_allow_html=True,
    )
    st.markdown("""
    <div class="feat-grid">
      <div class="feat-card">
        <div class="feat-badge">Mode 1</div>
        <div class="feat-icon" style="background:linear-gradient(135deg,#ede9fe,#ddd6fe);">🔍</div>
        <h3>Sheet Difference</h3>
        <p>Row-by-row comparison. Identify records present in one file but absent from the other, with full column context.</p>
      </div>
      <div class="feat-card">
        <div class="feat-badge">Mode 2</div>
        <div class="feat-icon" style="background:linear-gradient(135deg,#e0f2fe,#bae6fd);">📋</div>
        <h3>Stacked Comparison</h3>
        <p>Stack both files grouped by a control key. Paired records are highlighted; unmatched rows exported separately.</p>
      </div>
      <div class="feat-card">
        <div class="feat-badge">Mode 3</div>
        <div class="feat-icon" style="background:linear-gradient(135deg,#dcfce7,#bbf7d0);">🔢</div>
        <h3>Calculation Difference</h3>
        <p>Merge on a shared key and compute numeric deltas. Surfaces positive, negative, and zero variances instantly.</p>
      </div>
    </div>
    """, unsafe_allow_html=True)
    st.stop()


# ── Main Tabs ─────────────────────────────────────────────────────────────────
tab1, tab2, tab3 = st.tabs([
    "🔍  Sheet Difference",
    "📋  Stacked Comparison",
    "🔢  Calculation Difference",
])


# =============================================================================
# TAB 1 — SHEET DIFFERENCE
# =============================================================================
with tab1:
    sec("Column Mapping")
    ca, cb = st.columns(2, gap="large")
    with ca:
        st.caption(f"Columns from **{file_a.name}**")
        cols_a = st.multiselect(
            "cols_a", df_a.columns.tolist(),
            default=df_a.columns.tolist(), key="cols_a",
            label_visibility="collapsed",
        )
    with cb:
        st.caption(f"Columns from **{file_b.name}**")
        cols_b = st.multiselect(
            "cols_b", df_b.columns.tolist(),
            default=df_b.columns.tolist(), key="cols_b",
            label_visibility="collapsed",
        )

    st.write("")
    if len(cols_a) != len(cols_b):
        st.warning("Select the same number of columns from both sheets before running.")
    else:
        run_sd = st.button("Run Comparison", type="primary", key="run_sd")

        if run_sd:
            with st.spinner("Comparing sheets…"):
                df_a_sel = df_a[cols_a].copy()
                df_b_sel = df_b[cols_b].copy()
                df_b_sel.columns = df_a_sel.columns

                df_a_sel = df_a_sel.astype(str).apply(lambda x: x.str.strip()).drop_duplicates()
                df_b_sel = df_b_sel.astype(str).apply(lambda x: x.str.strip()).drop_duplicates()

                merged   = df_a_sel.merge(df_b_sel, how="outer", indicator=True)
                not_in_b = merged[merged["_merge"] == "left_only"].drop("_merge", axis=1)
                not_in_a = merged[merged["_merge"] == "right_only"].drop("_merge", axis=1)

                total_a   = len(df_a_sel)
                total_b   = len(df_b_sel)
                only_in_a = len(not_in_b)
                only_in_b = len(not_in_a)
                matched   = total_a - only_in_a

            st.write("")
            sec("Summary")
            metrics_row(
                mcard("Total in A", f"{total_a:,}", icon="📁", color="#9ca3af"),
                mcard("Total in B", f"{total_b:,}", icon="📁", color="#9ca3af"),
                mcard("Matched", f"{matched:,}", f"{matched/max(total_a,1)*100:.0f}% match rate", "#10b981", "✅"),
                mcard("Only in A", f"{only_in_a:,}", "missing from B", "#6366f1", "🔵"),
                mcard("Only in B", f"{only_in_b:,}", "missing from A", "#f59e0b", "🟡"),
            )

            output = io.BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                not_in_b.to_excel(writer, sheet_name="In_A_Not_in_B", index=False)
                not_in_a.to_excel(writer, sheet_name="In_B_Not_in_A", index=False)
            output.seek(0)

            st.download_button(
                "📥  Download Result",
                data=output,
                file_name="SheetDifference_Result.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

            st.write("")
            with st.expander(f"Rows in A not in B  ·  {only_in_a:,} records", expanded=only_in_a > 0):
                if not_in_b.empty:
                    st.success("Every row in A is present in B.")
                else:
                    st.dataframe(not_in_b, use_container_width=True, height=320)
            with st.expander(f"Rows in B not in A  ·  {only_in_b:,} records", expanded=only_in_b > 0):
                if not_in_a.empty:
                    st.success("Every row in B is present in A.")
                else:
                    st.dataframe(not_in_a, use_container_width=True, height=320)


# =============================================================================
# TAB 2 — STACKED COMPARISON
# =============================================================================
with tab2:
    sec("Configuration")
    c1, c2, c3 = st.columns([1, 1, 2], gap="large")
    with c1:
        name_a = st.text_input("Label for File A", "File A", placeholder="e.g. CDK", key="sc_na")
    with c2:
        name_b = st.text_input("Label for File B", "File B", placeholder="e.g. AXEL", key="sc_nb")
    with c3:
        control_col = st.text_input(
            "Control Column (case-insensitive match)",
            "Control number", key="sc_cc",
            help="Column used to pair records between files.",
        )

    st.write("")
    run_sc = st.button("Run Comparison", type="primary", key="run_sc")

    if run_sc:
        try:
            with st.spinner("Building stacked comparison…"):
                def find_col(df, target):
                    for c in df.columns:
                        if str(c).strip().lower() == target.strip().lower():
                            return c
                    raise KeyError(
                        f"Column '{target}' not found "
                        f"(case-insensitive). Available: {', '.join(str(x) for x in df.columns)}"
                    )

                ca_key = find_col(df_a, control_col)
                cb_key = find_col(df_b, control_col)
                df_a_w = df_a.rename(columns={ca_key: control_col}).copy()
                df_b_w = df_b.rename(columns={cb_key: control_col}).copy()

                df_a_w[control_col] = df_a_w[control_col].apply(
                    lambda x: str(x).strip() if pd.notna(x) else pd.NA)
                df_b_w[control_col] = df_b_w[control_col].apply(
                    lambda x: str(x).strip() if pd.notna(x) else pd.NA)

                df_a_w["Source"] = name_a
                df_b_w["Source"] = name_b

                all_cols = list(dict.fromkeys(
                    ["Source", control_col]
                    + [c for c in df_a_w.columns if c not in ("Source", control_col)]
                    + [c for c in df_b_w.columns if c not in ("Source", control_col)]
                ))
                df_a_w = df_a_w.reindex(columns=all_cols)
                df_b_w = df_b_w.reindex(columns=all_cols)

                a_set      = set(df_a_w[control_col].dropna().astype(str))
                b_set      = set(df_b_w[control_col].dropna().astype(str))
                paired_set = a_set & b_set
                a_only_set = a_set - b_set
                b_only_set = b_set - a_set

                stacked = pd.concat([df_a_w, df_b_w], ignore_index=True)

                cs = stacked[control_col].astype(str)
                stacked["PairStatus"] = ""
                stacked.loc[cs.isin(paired_set), "PairStatus"] = "Paired"
                stacked.loc[(stacked["Source"] == name_a) & cs.isin(a_only_set),
                            "PairStatus"] = f"{name_a}-only"
                stacked.loc[(stacked["Source"] == name_b) & cs.isin(b_only_set),
                            "PairStatus"] = f"{name_b}-only"

                stacked["_blk"] = cs.map(
                    lambda c: 0 if c in paired_set else (1 if c in a_only_set else 2))
                stacked["_ord"] = (stacked["Source"] != name_a).astype(int)
                stacked = stacked.sort_values(["_blk", control_col, "_ord"]).drop(
                    columns=["_blk", "_ord"]).reset_index(drop=True)

            st.write("")
            sec("Summary")
            metrics_row(
                mcard(f"Rows in {name_a}", f"{len(df_a_w):,}", icon="📁", color="#9ca3af"),
                mcard(f"Rows in {name_b}", f"{len(df_b_w):,}", icon="📁", color="#9ca3af"),
                mcard("Paired", f"{len(paired_set):,}",
                      f"{len(paired_set)/max(len(a_set),1)*100:.0f}% of A", "#10b981", "✅"),
                mcard(f"{name_a} only", f"{len(a_only_set):,}", "unmatched", "#6366f1", "🔵"),
                mcard(f"{name_b} only", f"{len(b_only_set):,}", "unmatched", "#f59e0b", "🟡"),
            )

            sh_a = sanitize_sheet_name(name_a, suffix="-only")
            sh_b = sanitize_sheet_name(name_b, suffix="-only")

            output = io.BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                stacked.to_excel(writer, sheet_name="Combined", index=False)
                df_a_w[df_a_w[control_col].isin(a_only_set)].to_excel(
                    writer, sheet_name=sh_a, index=False)
                df_b_w[df_b_w[control_col].isin(b_only_set)].to_excel(
                    writer, sheet_name=sh_b, index=False)
            output.seek(0)

            wb = load_workbook(output)
            ws = wb["Combined"]
            hdrs     = {cell.value: i + 1 for i, cell in enumerate(ws[1])}
            src_idx  = hdrs["Source"]
            ctrl_idx = hdrs[control_col]

            YELLOW = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
            BLUE   = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
            ws.auto_filter.ref = ws.dimensions
            ws.freeze_panes    = "A2"

            pu = {c.upper() for c in paired_set}
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                src  = (row[src_idx  - 1].value or "").strip().upper()
                ctrl = str(row[ctrl_idx - 1].value).strip().upper()
                if   src == name_b.upper():                fill = YELLOW
                elif src == name_a.upper() and ctrl in pu: fill = BLUE
                else:                                      continue
                for cell in row:
                    cell.fill = fill

            out_mem = io.BytesIO()
            wb.save(out_mem)
            out_mem.seek(0)

            st.download_button(
                "📥  Download Stacked Result",
                data=out_mem,
                file_name="StackedComparison_Result.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

            st.write("")
            with st.expander(f"Combined Preview  ·  top 50 of {len(stacked):,} rows", expanded=True):
                st.dataframe(stacked.head(50), use_container_width=True, height=380)

        except KeyError as e:
            st.error(str(e))


# =============================================================================
# TAB 3 — CALCULATION DIFFERENCE
# =============================================================================
with tab3:
    sec("Configuration")
    r1c1, r1c2 = st.columns(2, gap="large")
    with r1c1:
        name_a_c = st.text_input("Label for File A", "File A", placeholder="e.g. CDK", key="cd_na")
    with r1c2:
        name_b_c = st.text_input("Label for File B", "File B", placeholder="e.g. AXEL", key="cd_nb")

    r2c1, r2c2, r2c3 = st.columns(3, gap="large")
    with r2c1:
        ctrl_col = st.selectbox(
            "Key Column (common to both files)",
            df_a.columns, key="cd_key",
            help="Values in this column are used to match rows.",
        )
    with r2c2:
        num_a = st.selectbox(f"Numeric Column — {name_a_c}", df_a.columns, key="cd_na_col")
    with r2c3:
        num_b = st.selectbox(f"Numeric Column — {name_b_c}", df_b.columns, key="cd_nb_col")

    st.write("")
    run_cd = st.button("Run Comparison", type="primary", key="run_cd")

    if run_cd:
        if ctrl_col == num_a:
            st.error(f"Key column and numeric column for **{name_a_c}** must be different.")
        elif ctrl_col == num_b:
            st.error(f"Key column and numeric column for **{name_b_c}** must be different.")
        else:
            try:
                with st.spinner("Computing differences…"):
                    ta = df_a[[ctrl_col, num_a]].copy()
                    tb = df_b[[ctrl_col, num_b]].copy()

                    val_a = f"{num_a} [{name_a_c}]"
                    val_b = f"{num_b} [{name_b_c}]"
                    if val_a == val_b:
                        val_a, val_b = f"{val_a} (A)", f"{val_b} (B)"
                    ta = ta.rename(columns={num_a: val_a})
                    tb = tb.rename(columns={num_b: val_b})

                    ta[ctrl_col] = ta[ctrl_col].astype(str).str.strip()
                    tb[ctrl_col] = tb[ctrl_col].astype(str).str.strip()
                    ta[val_a]    = pd.to_numeric(ta[val_a], errors="coerce")
                    tb[val_b]    = pd.to_numeric(tb[val_b], errors="coerce")

                    merged = pd.merge(ta, tb, on=ctrl_col, how="inner")
                    merged["Difference"] = merged[val_a] - merged[val_b]

                match_pct = len(merged) / max(len(ta), 1) * 100
                ex_a      = len(ta) - len(merged)
                ex_b      = len(tb) - len(merged)
                mean_diff = merged["Difference"].mean() if not merged.empty else None

                st.write("")
                sec("Summary")
                metrics_row(
                    mcard(f"Rows in {name_a_c}", f"{len(ta):,}", icon="📁", color="#9ca3af"),
                    mcard(f"Rows in {name_b_c}", f"{len(tb):,}", icon="📁", color="#9ca3af"),
                    mcard("Matched", f"{len(merged):,}", f"{match_pct:.0f}% matched", "#10b981", "✅"),
                    mcard("A > B", f"{int((merged['Difference'] > 0).sum()):,}", "A exceeds B", "#6366f1", "📈"),
                    mcard("A < B", f"{int((merged['Difference'] < 0).sum()):,}", "B exceeds A", "#ef4444", "📉"),
                )

                if merged.empty:
                    st.warning(
                        "No matching rows found. "
                        "Verify the key column contains overlapping values in both files."
                    )
                else:
                    avg_str = f"{mean_diff:.2f}" if mean_diff is not None and pd.notna(mean_diff) else "N/A"
                    parts   = [f"**Average difference:** {avg_str}"]
                    if ex_a: parts.append(f"**{ex_a:,} rows from {name_a_c}** had no key match and were excluded")
                    if ex_b: parts.append(f"**{ex_b:,} rows from {name_b_c}** had no key match and were excluded")
                    st.info("  ·  ".join(parts))

                    output = io.BytesIO()
                    with pd.ExcelWriter(output, engine="openpyxl") as writer:
                        merged.to_excel(writer, sheet_name="Differences", index=False)
                    output.seek(0)

                    st.download_button(
                        "📥  Download Result",
                        data=output,
                        file_name="CalculationDifference_Result.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )

                    st.write("")
                    with st.expander(f"Differences Preview  ·  top 50 of {len(merged):,} rows", expanded=True):
                        st.dataframe(merged.head(50), use_container_width=True, height=380)

            except Exception as e:
                st.error(f"Error computing differences: {e}")
