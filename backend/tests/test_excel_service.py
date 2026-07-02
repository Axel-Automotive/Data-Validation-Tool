"""Unit tests for the comparison engine — the heart of the product."""
import io
import sys
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from app.services.excel_service import (  # noqa: E402
    run_sheet_difference, run_calc_difference, run_custom_rule, run_agg_compare,
    run_all_conditions, run_condition, get_result, _sanitize_sheet,
    _norm_series, _apply_filters,
)
from app.services import email_service  # noqa: E402
from fastapi import HTTPException  # noqa: E402


# ── Sheet difference ───────────────────────────────────────────────────────────

def test_sheet_diff_basic_match():
    a = pd.DataFrame({"K": ["1", "2", "3"]})
    b = pd.DataFrame({"K": ["1", "2", "4"]})
    r = run_sheet_difference(a, b, ["K"], ["K"])
    m = r["metrics"]
    assert m["matched"] == 2
    assert m["only_in_a"] == 1   # "3"
    assert m["only_in_b"] == 1   # "4"


def test_sheet_diff_int_vs_float_keys_match():
    # Same key stored as int in one file and float in the other must match.
    a = pd.DataFrame({"Stock": [1001, 1002, 1003]})
    b = pd.DataFrame({"Stock": [1001.0, 1002.0, 1003.0]})
    r = run_sheet_difference(a, b, ["Stock"], ["Stock"])
    assert r["metrics"]["matched"] == 3
    assert r["metrics"]["only_in_a"] == 0


def test_sheet_diff_unequal_columns_raises_400():
    a = pd.DataFrame({"K": [1], "V": [2]})
    with pytest.raises(HTTPException) as exc:
        run_sheet_difference(a, a, ["K"], ["K", "V"])
    assert exc.value.status_code == 400


def test_sheet_diff_column_named_merge_does_not_crash():
    # A compared column literally named "_merge" collided with the pandas
    # merge indicator and raised a 500 — must now work.
    a = pd.DataFrame({"_merge": ["1", "2", "3"]})
    b = pd.DataFrame({"_merge": ["1", "2", "4"]})
    r = run_sheet_difference(a, b, ["_merge"], ["_merge"])
    assert r["metrics"]["matched"] == 2
    assert r["metrics"]["only_in_a"] == 1


# ── Calc difference ─────────────────────────────────────────────────────────────

def test_calc_diff_no_fanout_on_duplicate_keys():
    # Duplicate keys must NOT produce an N×M cartesian explosion.
    a = pd.DataFrame({"K": ["1", "1", "2"], "V": [10, 20, 30]})
    b = pd.DataFrame({"K": ["1", "1", "2"], "V": [5, 5, 5]})
    r = run_calc_difference(a, b, "AXEL", "DMS", "K", "V", "V")
    assert r["metrics"]["matched"] == 2          # not 5
    assert r["metrics"]["excluded_a"] >= 0       # never negative


def test_calc_diff_reports_duplicate_keys():
    # AXEL key "1" appears 3× — the drops must be reported, not silent.
    a = pd.DataFrame({"K": ["1", "1", "1", "2"], "V": [10, 20, 30, 40]})
    b = pd.DataFrame({"K": ["1", "2"], "V": [10, 40]})
    r = run_calc_difference(a, b, "AXEL", "DMS", "K", "V", "V")
    assert r["metrics"]["duplicate_keys_a"] == 1     # one key ("1") had dupes
    assert r["metrics"]["duplicate_keys_b"] == 0
    dup = r["_frames"]["Duplicate Keys AXEL"]
    assert dup[dup["K"] == "1"]["Occurrences"].iloc[0] == 3


def test_calc_diff_difference_value():
    a = pd.DataFrame({"K": ["x"], "V": [100]})
    b = pd.DataFrame({"K": ["x"], "V": [70]})
    r = run_calc_difference(a, b, "AXEL", "DMS", "K", "V", "V")
    assert r["metrics"]["matched"] == 1
    assert r["metrics"]["a_gt_b"] == 1


def test_calc_diff_exports_unmatched_breaks():
    # Key "2" is only in AXEL, "3" only in DMS — both are reconciliation breaks
    # that must appear in the output, not just the metric counts.
    a = pd.DataFrame({"K": ["1", "2"], "V": [100, 50]})
    b = pd.DataFrame({"K": ["1", "3"], "V": [100, 70]})
    r = run_calc_difference(a, b, "AXEL", "DMS", "K", "V", "V")
    m = r["metrics"]
    assert m["matched"] == 1
    assert m["unmatched_a"] == 1 and m["unmatched_b"] == 1
    frames = r["_frames"]
    assert "Unmatched AXEL" in frames and "Unmatched DMS" in frames
    assert frames["Unmatched AXEL"]["K"].tolist() == ["2"]
    assert frames["Unmatched DMS"]["K"].tolist() == ["3"]
    # And they ride into the on-screen preview too.
    assert r["preview"]["unmatched_a"]["total"] == 1
    assert r["preview"]["unmatched_b"]["total"] == 1


# ── Custom rule ─────────────────────────────────────────────────────────────────

def test_custom_rule_numeric_and_text():
    a = pd.DataFrame({"Deal": ["1", "2", "3"], "G": [100, 200, 300], "S": ["O", "C", "O"]})
    b = pd.DataFrame({"DealNo": ["1", "2", "3"], "R": [100, 205, 300], "S": ["O", "C", "C"]})
    cfg = {
        "key_axel": "Deal", "key_dms": "DealNo",
        "checks": [
            {"axel_col": "G", "dms_col": "R", "mode": "numeric", "op": "eq", "tolerance": 10},
            {"axel_col": "S", "dms_col": "S", "mode": "text", "op": "eq"},
        ],
    }
    r = run_custom_rule(a, b, cfg)
    m = r["metrics"]
    assert m["matched"] == 3
    assert m["passed"] == 2     # deal 3 fails the text check (O vs C)
    assert m["failed"] == 1


def test_composite_key_calc_and_rule():
    # Row identity is Acct+Deal; 206/1 vs 206/9 is a break on each side.
    a = pd.DataFrame({"Acct": ["205", "205", "206"], "Deal": ["1", "2", "1"], "V": [100, 200, 300]})
    b = pd.DataFrame({"Acct": ["205", "205", "206"], "Deal": ["1", "2", "9"], "V": [100, 250, 300]})

    r = run_calc_difference(a, b, "AXEL", "DMS", ["Acct", "Deal"], "V", "V", ["Acct", "Deal"])
    m = r["metrics"]
    assert m["matched"] == 2 and m["unmatched_a"] == 1 and m["unmatched_b"] == 1
    assert "Acct + Deal" in r["_frames"]["Differences"].columns

    cfg = {"key_axel": ["Acct", "Deal"],
           "checks": [{"axel_col": "V", "dms_col": "V", "mode": "numeric", "op": "eq", "tolerance": 0}]}
    rr = run_custom_rule(a, b, cfg)
    assert rr["metrics"]["matched"] == 2 and rr["metrics"]["passed"] == 1

    # Single-string key stays fully backward compatible.
    assert run_calc_difference(a, b, "AXEL", "DMS", "Acct", "V", "V")["metrics"]["matched"] == 2


def test_calc_diff_on_duplicate_sum():
    # AXEL has two journal lines for deal 1 (15 + 5); DMS carries the total 20.
    a = pd.DataFrame({"K": ["1", "1", "2"], "V": [15, 5, 30]})
    b = pd.DataFrame({"K": ["1", "2"], "V": [20, 30]})
    # Default "first" keeps only 15 → a false break of -5.
    first = run_calc_difference(a, b, "AXEL", "DMS", "K", "V", "V")
    assert first["metrics"]["matched"] == 2
    # Summing duplicates reconciles deal 1 exactly.
    summed = run_calc_difference(a, b, "AXEL", "DMS", "K", "V", "V", on_duplicate="sum")
    diffs = summed["_frames"]["Differences"]
    assert diffs[diffs["K"] == "1"]["Difference"].iloc[0] == 0
    assert summed["metrics"]["zero"] == 2


def test_key_canonicalization_matches_across_formats():
    # Same deals, keys differ by leading zeros / case / date format.
    a = pd.DataFrame({"K": ["007", "abc"], "V": [100, 200]})
    b = pd.DataFrame({"K": ["7", "ABC"], "V": [100, 200]})
    # Without canonicalisation nothing matches.
    assert run_calc_difference(a, b, "AXEL", "DMS", "K", "V", "V")["metrics"]["matched"] == 0
    # With strip_zeros + uppercase, both match.
    r = run_calc_difference(a, b, "AXEL", "DMS", "K", "V", "V",
                            key_norm={"strip_zeros": True, "uppercase": True})
    assert r["metrics"]["matched"] == 2

    # Date-format normalisation on the key.
    da = pd.DataFrame({"D": ["6/27/26"], "V": [10]})
    db = pd.DataFrame({"D": ["2026-06-27"], "V": [10]})
    assert run_calc_difference(da, db, "AXEL", "DMS", "D", "V", "V")["metrics"]["matched"] == 0
    r2 = run_calc_difference(da, db, "AXEL", "DMS", "D", "V", "V", key_norm={"parse_date": True})
    assert r2["metrics"]["matched"] == 1


def test_agg_compare_sum_by_group():
    # AXEL has detail lines; DMS carries per-account totals. Reconcile the sums.
    a = pd.DataFrame({"Acct": ["205", "205", "206"], "Amt": [15420.51, 18946.75, 300.0]})
    b = pd.DataFrame({"Acct": ["205", "206"], "Total": [34367.26, 305.0]})
    cfg = {"group_axel": "Acct", "group_dms": "Acct", "metric": "sum",
           "value_axel": "Amt", "value_dms": "Total", "tolerance": 0}
    r = run_agg_compare(a, b, cfg)
    m = r["metrics"]
    assert m["matched"] == 2
    assert m["passed"] == 1     # 205 sums to 34367.26 (pass), 206 off by 5 (fail)
    assert m["failed"] == 1
    agg = r["_frames"]["Aggregate"]
    assert agg[agg["Acct"] == "205"]["Result"].iloc[0] == "Pass"
    assert agg[agg["Acct"] == "206"]["Result"].iloc[0] == "Fail"


def test_agg_compare_count_needs_no_value():
    a = pd.DataFrame({"Acct": ["205", "205", "206"]})
    b = pd.DataFrame({"Acct": ["205", "205", "206", "206"]})
    r = run_agg_compare(a, b, {"group_axel": "Acct", "metric": "count", "tolerance": 0})
    agg = r["_frames"]["Aggregate"]
    assert agg[agg["Acct"] == "205"]["Result"].iloc[0] == "Pass"   # 2 vs 2
    assert agg[agg["Acct"] == "206"]["Result"].iloc[0] == "Fail"   # 1 vs 2


def test_agg_compare_percentage_tolerance():
    a = pd.DataFrame({"G": ["x"], "V": [1000.0]})
    b = pd.DataFrame({"G": ["x"], "V": [1005.0]})
    # 5 diff on 1005 ≈ 0.5% → passes a 1% band, fails absolute-0.
    strict = run_agg_compare(a, b, {"group_axel": "G", "metric": "sum",
                                    "value_axel": "V", "value_dms": "V", "tolerance": 0})
    assert strict["metrics"]["passed"] == 0
    lax = run_agg_compare(a, b, {"group_axel": "G", "metric": "sum",
                                 "value_axel": "V", "value_dms": "V", "tolerance_pct": 1.0})
    assert lax["metrics"]["passed"] == 1


def test_custom_rule_percentage_tolerance():
    # 3 diff on 1000 = 0.3% → passes a 1% band; fails absolute-0.
    a = pd.DataFrame({"K": ["1"], "V": [1003]})
    b = pd.DataFrame({"K": ["1"], "V": [1000]})
    strict = {"key_axel": "K", "checks": [
        {"axel_col": "V", "dms_col": "V", "mode": "numeric", "op": "eq", "tolerance": 0}]}
    assert run_custom_rule(a, b, strict)["metrics"]["passed"] == 0
    pct = {"key_axel": "K", "checks": [
        {"axel_col": "V", "dms_col": "V", "mode": "numeric", "op": "eq", "tolerance_pct": 1.0}]}
    assert run_custom_rule(a, b, pct)["metrics"]["passed"] == 1


def test_custom_rule_warning_severity_does_not_fail_row():
    # The value check is an ERROR (passes); the status check is a WARNING (fails).
    a = pd.DataFrame({"K": ["1"], "V": [100], "S": ["Open"]})
    b = pd.DataFrame({"K": ["1"], "V": [100], "S": ["Closed"]})
    cfg = {"key_axel": "K", "checks": [
        {"axel_col": "V", "dms_col": "V", "mode": "numeric", "op": "eq", "tolerance": 0, "severity": "error"},
        {"axel_col": "S", "dms_col": "S", "mode": "text", "op": "eq", "severity": "warning"}]}
    m = run_custom_rule(a, b, cfg)["metrics"]
    assert m["passed"] == 1        # row still passes (only a warning tripped)
    assert m["failed"] == 0
    assert m["warnings"] == 1


def test_fuzzy_key_matching_pairs_near_misses():
    # VINs differ by one char; exact matching misses them, fuzzy pairs them.
    a = pd.DataFrame({"VIN": ["3VVSX7B21RM066810", "JF2SJAEC7JH616247"], "V": [100, 200]})
    b = pd.DataFrame({"VIN": ["3VVSX7B21RM066811", "JF2SJAEC7JH616247"], "V": [100, 200]})
    exact = run_calc_difference(a, b, "AXEL", "DMS", "VIN", "V", "V")
    assert exact["metrics"]["matched"] == 1        # only the identical VIN
    fuzzy = run_calc_difference(a, b, "AXEL", "DMS", "VIN", "V", "V",
                                fuzzy={"enabled": True, "threshold": 0.9})
    assert fuzzy["metrics"]["matched"] == 2        # near-miss VIN now paired


def test_custom_rule_requires_key_and_checks():
    a = pd.DataFrame({"K": ["1"]})
    with pytest.raises(HTTPException):
        run_custom_rule(a, a, {"key_axel": "", "checks": []})


def test_custom_rule_exports_unmatched_breaks():
    # Deal "3" only in AXEL, "4" only in DMS → must be exported as breaks.
    a = pd.DataFrame({"Deal": ["1", "2", "3"], "G": [100, 200, 300]})
    b = pd.DataFrame({"DealNo": ["1", "2", "4"], "R": [100, 200, 400]})
    cfg = {"key_axel": "Deal", "key_dms": "DealNo", "checks": [
        {"axel_col": "G", "dms_col": "R", "mode": "numeric", "op": "eq", "tolerance": 0}]}
    r = run_custom_rule(a, b, cfg)
    m = r["metrics"]
    assert m["matched"] == 2
    assert m["unmatched_a"] == 1 and m["unmatched_b"] == 1
    frames = r["_frames"]
    assert frames["Unmatched AXEL"]["Deal"].tolist() == ["3"]
    assert frames["Unmatched DMS"]["DealNo"].tolist() == ["4"]


def test_custom_rule_duplicate_keys_do_not_inflate_unmatched():
    # AXEL has key "1" twice; both sides share keys 1 and 2. After dedup there
    # are 2 distinct matched keys and 0 unmatched — the raw row count must not
    # leak into unmatched_a as a phantom unmatched row.
    a = pd.DataFrame({"K": ["1", "1", "2"], "V": [100, 100, 200]})
    b = pd.DataFrame({"K": ["1", "2"],      "V": [100, 200]})
    cfg = {"key_axel": "K", "checks": [
        {"axel_col": "V", "dms_col": "V", "mode": "numeric", "op": "eq", "tolerance": 0}]}
    m = run_custom_rule(a, b, cfg)["metrics"]
    assert m["matched"] == 2
    assert m["unmatched_a"] == 0
    assert m["unmatched_b"] == 0


# ── Helpers ──────────────────────────────────────────────────────────────────────

def test_sanitize_sheet_strips_newlines():
    assert "\n" not in _sanitize_sheet("Stock\nNo.")
    assert _sanitize_sheet("a/b:c*[d]") and "/" not in _sanitize_sheet("a/b:c*[d]")
    assert _sanitize_sheet("") == "Sheet"


def test_norm_series_handles_float_and_nan():
    s = pd.Series([1001.0, 1002.0, None])
    out = list(_norm_series(s, na=""))
    assert out[0] == "1001" and out[1] == "1002" and out[2] == ""


# ── Combined report ──────────────────────────────────────────────────────────────

def test_run_all_combined_report_full_rows_numeric():
    n = 150
    a = pd.DataFrame({"Deal": [str(i) for i in range(n)], "G": [i for i in range(n)]})
    b = pd.DataFrame({"Deal": [str(i) for i in range(n)], "R": [i for i in range(n)]})
    conds = [{
        "id": "c1", "name": "Rule", "type": "custom_rule", "enabled": True,
        "config": {"key_axel": "Deal", "checks": [
            {"axel_col": "G", "dms_col": "R", "mode": "numeric", "op": "eq", "tolerance": 0}]},
    }]
    rid, results = run_all_conditions(conds, a, b)
    assert results[0]["metrics"]["matched"] == n
    data, _ = get_result(rid)
    wb = load_workbook(io.BytesIO(data))
    ws = next(s for s in wb.worksheets if "Rule" in s.title)
    assert ws.max_row - 1 == n                       # not capped at 100
    assert isinstance(ws.cell(row=2, column=2).value, (int, float))


def test_run_all_has_rag_health_and_run_info():
    a = pd.DataFrame({"K": ["1", "2", "3"], "V": [100, 200, 300]})
    b = pd.DataFrame({"K": ["1", "2", "3"], "V": [100, 200, 300]})   # perfect match
    conds = [{"id": "c1", "name": "Vals", "type": "calc_diff", "enabled": True,
              "config": {"key_axel": "K", "val_axel": "V", "val_dms": "V"}}]
    rid, _ = run_all_conditions(conds, a, b, {"axel_name": "a.xlsx", "dms_name": "b.csv"})
    wb = load_workbook(io.BytesIO(get_result(rid)[0]))
    assert "Run Info" in wb.sheetnames
    ws = wb["Summary"]
    hdr = [c.value for c in ws[1]]
    assert "Health" in hdr
    health = ws.cell(row=2, column=hdr.index("Health") + 1).value
    assert health == "✓ Pass"                         # 100% match → green/Pass
    # Totals row present.
    assert ws.cell(row=ws.max_row, column=2).value == "TOTALS"


# ── Row filtering ──────────────────────────────────────────────────────────────

def test_apply_filters_eq_and_numeric():
    df = pd.DataFrame({"Status": ["Active", "Closed", "Active"], "Amount": [10, -5, 20]})
    out = _apply_filters(df, [{"col": "Status", "op": "eq", "value": "Active"}], None)
    assert list(out["Amount"]) == [10, 20]
    out = _apply_filters(df, [{"col": "Amount", "op": "gt", "value": 0}], None)
    assert list(out["Status"]) == ["Active", "Active"]


def test_apply_filters_in_and_blank():
    df = pd.DataFrame({"Branch": ["North", "South", "", "North"]})
    out = _apply_filters(df, [{"col": "Branch", "op": "in", "value": ["North"]}], None)
    assert len(out) == 2
    out = _apply_filters(df, [{"col": "Branch", "op": "not_blank", "value": ""}], None)
    assert len(out) == 3


def test_apply_filters_row_range_is_1_based_inclusive():
    df = pd.DataFrame({"K": list(range(1, 11))})        # 1..10
    out = _apply_filters(df, None, {"start": 2, "end": 5})
    assert list(out["K"]) == [2, 3, 4, 5]


def test_run_condition_applies_filters_before_compare():
    # Only "Active" rows should be compared, so the inactive mismatch is ignored.
    a = pd.DataFrame({"K": ["1", "2", "3"], "S": ["Active", "Active", "Closed"]})
    b = pd.DataFrame({"K": ["1", "2", "9"], "S": ["Active", "Active", "Closed"]})
    cfg = {"col_pairs": [{"axel": "K", "dms": "K"}],
           "filters": {"axel": [{"col": "S", "op": "eq", "value": "Active"}],
                       "dms":  [{"col": "S", "op": "eq", "value": "Active"}]}}
    r = run_condition(a, b, "sheet_diff", cfg)
    assert r["metrics"]["matched"] == 2
    assert r["metrics"]["only_in_a"] == 0      # "3"/"9" filtered out


# ── Failing-cell highlighting ────────────────────────────────────────────────────

def test_custom_rule_highlights_failing_cells():
    a = pd.DataFrame({"Deal": ["1", "2"], "G": [100, 200]})
    b = pd.DataFrame({"Deal": ["1", "2"], "R": [100, 999]})   # deal 2 fails
    cfg = {"key_axel": "Deal", "checks": [
        {"axel_col": "G", "dms_col": "R", "mode": "numeric", "op": "eq", "tolerance": 0}]}
    r = run_custom_rule(a, b, cfg)
    assert "Failing Columns" in r["preview"]["results"]["columns"]
    data, _ = get_result(r["result_id"])
    wb = load_workbook(io.BytesIO(data))
    ws = wb["Rule_Results"]
    # The failing row's source cells must be red-filled; the passing row must not.
    hdr = {c.value: i + 1 for i, c in enumerate(ws[1])}
    def filled(row, name):
        return ws.cell(row=row, column=hdr[name]).fill.fgColor.rgb in ("FFFECACA", "00FECACA")
    assert filled(3, "G [AXEL]") and filled(3, "R [DMS]")     # row 3 = deal 2
    assert not filled(2, "G [AXEL]")                          # row 2 = deal 1 passed


# ── Email helpers ────────────────────────────────────────────────────────────────

def test_clean_recipients_dedupes_and_validates():
    assert email_service.clean_recipients(["A@x.com", "a@x.com", " b@y.com "]) == ["A@x.com", "b@y.com"]
    with pytest.raises(HTTPException):
        email_service.clean_recipients(["notanemail"])
