"""All Excel comparison logic — ad-hoc and condition-based."""
from __future__ import annotations
import io
import json
import re
import time
import uuid
from pathlib import Path

import pandas as pd
from fastapi import HTTPException
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── Result store (persisted to disk so download links survive restarts) ────────
RESULTS_DIR = Path(__file__).parent.parent.parent / "data" / "results"


def get_result(result_id: str) -> tuple[bytes, str] | None:
    # Guard against path traversal — result_id must be a bare uuid-like token.
    if not re.fullmatch(r"[A-Za-z0-9\-]+", result_id or ""):
        return None
    data_p = RESULTS_DIR / f"{result_id}.xlsx"
    meta_p = RESULTS_DIR / f"{result_id}.json"
    if not data_p.exists():
        return None
    filename = "Report.xlsx"
    try:
        filename = json.loads(meta_p.read_text()).get("filename", filename)
    except Exception:
        pass
    return data_p.read_bytes(), filename


def _save_result(data: bytes, filename: str) -> str:
    rid = str(uuid.uuid4())
    RESULTS_DIR.mkdir(parents=True, exist_ok=True)
    (RESULTS_DIR / f"{rid}.xlsx").write_bytes(data)
    (RESULTS_DIR / f"{rid}.json").write_text(json.dumps({"filename": filename}))
    return rid


def cleanup_old_results(max_age_days: int = 14) -> int:
    """Delete result files older than max_age_days. Returns count removed."""
    if not RESULTS_DIR.exists():
        return 0
    cutoff = time.time() - max_age_days * 86400
    removed = 0
    for p in RESULTS_DIR.glob("*"):
        try:
            if p.stat().st_mtime < cutoff:
                p.unlink()
                removed += 1
        except OSError:
            pass
    return removed


# ── Utilities ─────────────────────────────────────────────────────────────────

def _to_records(df: pd.DataFrame, limit: int = 100) -> list[dict]:
    return df.head(limit).astype(str).to_dict(orient="records")


def _sanitize_sheet(name: str, suffix: str = "", max_len: int = 31) -> str:
    # Excel forbids  / \ ? * [ ] :  and control chars (newline/tab) in titles,
    # and a title cannot start/end with an apostrophe or be blank.
    name = re.sub(r'[/\\?*\[\]:\n\r\t]', " ", str(name))
    name = re.sub(r"\s+", " ", name).strip().strip("'").strip() or "Sheet"
    return name[: max_len - len(suffix)].strip() + suffix


def _find_col(df: pd.DataFrame, target: str) -> str:
    for c in df.columns:
        if str(c).strip().lower() == target.strip().lower():
            return c
    avail = ", ".join(str(c) for c in df.columns)
    raise HTTPException(400, f"Column '{target}' not found. Available: {avail}")


def _norm_series(s: pd.Series, na=""):
    """Normalise a column for comparison: trim text and render whole-number
    floats without a trailing '.0', so the same key stored as int 1001 in one
    file and float 1001.0 in the other still matches. NaN → `na`."""
    num = pd.to_numeric(s, errors="coerce")
    if s.notna().any() and num[s.notna()].notna().all():
        def _fmt(v):
            if pd.isna(v):
                return na
            f = float(v)
            return str(int(f)) if f.is_integer() else repr(f)
        return num.map(_fmt)
    return s.map(lambda v: na if pd.isna(v) else str(v).strip())


def _dup_keys(s: pd.Series, key_name: str) -> pd.DataFrame:
    """Keys that occur more than once in `s`, with their occurrence count.

    Calc/rule comparisons keep only the first row per key, so duplicates are
    silently dropped; this makes that data loss visible instead."""
    vc = s.value_counts()
    vc = vc[vc > 1]
    return pd.DataFrame({key_name: vc.index.astype(str), "Occurrences": vc.values.astype(int)})


# ── Join keys (single column or composite of several) ──────────────────────────

_KEY_SEP = "␟"   # unit separator — joins composite key parts, unlikely in data


def _key_cols(spec) -> list[str]:
    """A key spec is a single column name (str) or a list of names. Normalise
    to a list of non-blank names. Keeps single-string configs working as-is."""
    if isinstance(spec, (list, tuple)):
        return [str(c).strip() for c in spec if str(c).strip()]
    s = str(spec or "").strip()
    return [s] if s else []


def _key_label(cols: list[str]) -> str:
    """Display name for a (possibly composite) key."""
    return " + ".join(cols)


def _canon_value(s: pd.Series, opts: dict | None) -> pd.Series:
    """Optional key canonicalisation to stop trivially-different keys from being
    reported as breaks. `opts` flags (all optional):
      parse_date   → parse to a date and render ISO (YYYY-MM-DD)
      strip_zeros  → drop leading zeros ("007" → "7", "0" stays "0")
      uppercase    → upper-case
      alnum_only   → keep only letters/digits (drops spaces, dashes, punctuation)
    """
    if not opts:
        return s
    out = s
    if opts.get("parse_date"):
        d = pd.to_datetime(out, errors="coerce")
        out = out.where(d.isna(), d.dt.strftime("%Y-%m-%d"))
    if opts.get("strip_zeros"):
        out = out.str.replace(r"^0+(?=.)", "", regex=True)
    if opts.get("uppercase"):
        out = out.str.upper()
    if opts.get("alnum_only"):
        out = out.str.replace(r"[^A-Za-z0-9]", "", regex=True)
    return out


def _composite_key(df: pd.DataFrame, cols: list[str], na="", norm: dict | None = None) -> pd.Series:
    """Build a normalised key Series from one or more columns of `df`.
    Each part is normalised via `_norm_series`, optionally canonicalised
    (`norm`), then joined with `_KEY_SEP`."""
    if not cols:
        raise HTTPException(400, "A join key column is required.")
    parts = [_canon_value(_norm_series(df[_find_col(df, c)], na=na), norm) for c in cols]
    key = parts[0]
    for p in parts[1:]:
        key = key.str.cat(p, sep=_KEY_SEP)
    return key


# ── Row filtering (applied to either file before any comparison) ────────────────
#
# config["filters"] = {
#   "axel": [ { "col", "op", "value" } ],   # combined with AND
#   "dms":  [ { "col", "op", "value" } ],
#   "row_range_axel": { "start": 2, "end": 500 },   # 1-based, inclusive, optional
#   "row_range_dms":  { ... },
# }

def _apply_one_filter(df: pd.DataFrame, f: dict) -> pd.DataFrame:
    col = _find_col(df, f["col"])
    op  = f.get("op", "eq")
    val = f.get("value", "")
    s   = df[col]

    if op in ("gt", "lt", "gte", "lte"):
        num = pd.to_numeric(s, errors="coerce")
        try:
            v = float(val)
        except (TypeError, ValueError):
            raise HTTPException(400, f"Filter on '{f['col']}' needs a number for '{op}'.")
        mask = {"gt": num > v, "lt": num < v, "gte": num >= v, "lte": num <= v}[op]
        return df[mask.fillna(False)]

    if op in ("eq", "ne"):
        norm   = _norm_series(s)
        target = _norm_series(pd.Series([val])).iloc[0]
        mask   = norm == target
        return df[mask if op == "eq" else ~mask]

    if op in ("in", "not_in"):
        vals = val if isinstance(val, list) else [v.strip() for v in str(val).split(",")]
        targets = set(_norm_series(pd.Series(vals)))
        mask = _norm_series(s).isin(targets)
        return df[mask if op == "in" else ~mask]

    if op in ("contains", "not_contains"):
        mask = s.astype(str).str.contains(str(val), case=False, na=False, regex=False)
        return df[mask if op == "contains" else ~mask]

    if op in ("is_blank", "not_blank"):
        blank = s.isna() | (s.astype(str).str.strip() == "")
        return df[blank if op == "is_blank" else ~blank]

    raise HTTPException(400, f"Unknown filter operator: {op}")


def _apply_filters(df: pd.DataFrame, filters: list[dict] | None,
                   row_range: dict | None) -> pd.DataFrame:
    out = df
    for f in filters or []:
        if f.get("col"):
            out = _apply_one_filter(out, f)
    if row_range:
        start = row_range.get("start")
        end   = row_range.get("end")
        lo = (int(start) - 1) if start else 0           # 1-based → 0-based
        hi = int(end) if end else len(out)              # inclusive end
        out = out.iloc[max(lo, 0):hi]
    return out


# ── Sheet Difference ───────────────────────────────────────────────────────────

def run_sheet_difference(
    df_a: pd.DataFrame,
    df_b: pd.DataFrame,
    cols_a: list[str],
    cols_b: list[str],
) -> dict:
    if len(cols_a) != len(cols_b):
        raise HTTPException(400, f"Select an equal number of columns: {len(cols_a)} vs {len(cols_b)}.")
    if not cols_a:
        raise HTTPException(400, "Select at least one column to compare.")
    cols_a = [_find_col(df_a, c) for c in cols_a]
    cols_b = [_find_col(df_b, c) for c in cols_b]
    sel_a = df_a[cols_a].copy()
    sel_b = df_b[cols_b].copy()
    sel_b.columns = sel_a.columns          # normalise column names

    sel_a = sel_a.apply(_norm_series).drop_duplicates()
    sel_b = sel_b.apply(_norm_series).drop_duplicates()

    # Use an indicator column name that can't collide with a compared column
    # (pandas raises if we ask for indicator=True and "_merge" already exists).
    ind = "_merge"
    while ind in sel_a.columns:
        ind += "_"
    merged   = sel_a.merge(sel_b, how="outer", indicator=ind)
    not_in_b = merged[merged[ind] == "left_only"].drop(ind, axis=1)
    not_in_a = merged[merged[ind] == "right_only"].drop(ind, axis=1)

    total_a = len(sel_a)
    total_b = len(sel_b)
    only_a  = len(not_in_b)
    only_b  = len(not_in_a)
    matched = total_a - only_a

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        not_in_b.to_excel(w, sheet_name="In_A_Not_in_B", index=False)
        not_in_a.to_excel(w, sheet_name="In_B_Not_in_A", index=False)
    buf.seek(0)

    return {
        "type": "sheet_diff",
        "metrics": {
            "total_a": total_a, "total_b": total_b,
            "matched": matched, "only_in_a": only_a, "only_in_b": only_b,
            "match_rate": round(matched / max(total_a, 1) * 100, 1),
        },
        "preview": {
            "not_in_b": {"data": _to_records(not_in_b), "total": only_a,  "columns": not_in_b.columns.tolist()},
            "not_in_a": {"data": _to_records(not_in_a), "total": only_b,  "columns": not_in_a.columns.tolist()},
        },
        "result_id": _save_result(buf.getvalue(), "SheetDifference_Result.xlsx"),
        "_frames": {"In_A_Not_in_B": not_in_b, "In_B_Not_in_A": not_in_a},
    }


# ── Stacked Comparison ────────────────────────────────────────────────────────

def run_stacked_comparison(
    df_a: pd.DataFrame,
    df_b: pd.DataFrame,
    name_a: str,
    name_b: str,
    control_col_a: str,
    control_col_b: str | None = None,
) -> dict:
    if control_col_b is None:
        control_col_b = control_col_a

    a_key = _find_col(df_a, control_col_a)
    b_key = _find_col(df_b, control_col_b)
    CTRL = "ControlKey"      # unified name in the stacked frame

    dfa = df_a.rename(columns={a_key: CTRL}).copy()
    dfb = df_b.rename(columns={b_key: CTRL}).copy()

    for df_ in (dfa, dfb):
        df_[CTRL] = _norm_series(df_[CTRL], na=pd.NA)

    dfa["Source"] = name_a
    dfb["Source"] = name_b

    all_cols = list(dict.fromkeys(
        ["Source", CTRL]
        + [c for c in dfa.columns if c not in ("Source", CTRL)]
        + [c for c in dfb.columns if c not in ("Source", CTRL)]
    ))
    dfa = dfa.reindex(columns=all_cols)
    dfb = dfb.reindex(columns=all_cols)

    a_set   = set(dfa[CTRL].dropna().astype(str))
    b_set   = set(dfb[CTRL].dropna().astype(str))
    paired  = a_set & b_set
    a_only  = a_set - b_set
    b_only  = b_set - a_set

    stacked = pd.concat([dfa, dfb], ignore_index=True)
    cs      = stacked[CTRL].astype(str)
    stacked["PairStatus"] = ""
    stacked.loc[cs.isin(paired), "PairStatus"] = "Paired"
    stacked.loc[(stacked["Source"] == name_a) & cs.isin(a_only), "PairStatus"] = f"{name_a}-only"
    stacked.loc[(stacked["Source"] == name_b) & cs.isin(b_only), "PairStatus"] = f"{name_b}-only"

    stacked["_blk"] = cs.map(lambda c: 0 if c in paired else (1 if c in a_only else 2))
    stacked["_ord"] = (stacked["Source"] != name_a).astype(int)
    stacked = stacked.sort_values(["_blk", CTRL, "_ord"]).drop(
        columns=["_blk", "_ord"]).reset_index(drop=True)

    # Rename CTRL back to the original name for export
    stacked = stacked.rename(columns={CTRL: control_col_a})

    sh_a = _sanitize_sheet(name_a, suffix="-only")
    sh_b = _sanitize_sheet(name_b, suffix="-only")

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        stacked.to_excel(w, sheet_name="Combined", index=False)
        dfa.rename(columns={CTRL: control_col_a})[
            dfa[CTRL].isin(a_only)].to_excel(w, sheet_name=sh_a, index=False)
        dfb.rename(columns={CTRL: control_col_a})[
            dfb[CTRL].isin(b_only)].to_excel(w, sheet_name=sh_b, index=False)
    buf.seek(0)

    # Colour the Combined sheet
    wb = load_workbook(buf)
    ws = wb["Combined"]
    hdrs    = {cell.value: i + 1 for i, cell in enumerate(ws[1])}
    src_idx = hdrs.get("Source", 1)
    YELLOW  = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    BLUE    = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes    = "A2"
    pu = {c.upper() for c in paired}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        src  = (row[src_idx - 1].value or "").strip().upper()
        ctrl_val = str(row[hdrs.get(control_col_a, 2) - 1].value).strip().upper()
        if   src == name_b.upper():                         fill = YELLOW
        elif src == name_a.upper() and ctrl_val in pu:      fill = BLUE
        else:                                               continue
        for cell in row: cell.fill = fill

    out = io.BytesIO(); wb.save(out); out.seek(0)

    return {
        "type": "stacked",
        "metrics": {
            "rows_a": len(dfa), "rows_b": len(dfb),
            "paired": len(paired), "a_only": len(a_only), "b_only": len(b_only),
            "pair_rate": round(len(paired) / max(len(a_set), 1) * 100, 1),
        },
        "preview": {
            "combined": {"data": _to_records(stacked), "total": len(stacked),
                         "columns": stacked.columns.tolist()},
        },
        "result_id": _save_result(out.getvalue(), "StackedComparison_Result.xlsx"),
        "_frames": {"Combined": stacked},
    }


# ── Calculation Difference ────────────────────────────────────────────────────

def run_calc_difference(
    df_a: pd.DataFrame,
    df_b: pd.DataFrame,
    name_a: str,
    name_b: str,
    key_col_a: str,
    num_col_a: str,
    num_col_b: str,
    key_col_b: str | None = None,
    key_norm: dict | None = None,
    on_duplicate: str = "first",
) -> dict:
    # Key can be a single column or a composite (list of columns). The DMS side
    # mirrors AXEL when its key is blank.
    key_cols_a = _key_cols(key_col_a)
    key_cols_b = _key_cols(key_col_b) or key_cols_a
    key_disp_a = _key_label(key_cols_a)
    key_disp_b = _key_label(key_cols_b)
    num_col_a = _find_col(df_a, num_col_a)
    num_col_b = _find_col(df_b, num_col_b)

    val_a = f"{num_col_a} [{name_a}]"
    val_b = f"{num_col_b} [{name_b}]"
    if val_a == val_b:
        val_a, val_b = f"{val_a} (A)", f"{val_b} (B)"

    ta = pd.DataFrame({"KEY": _composite_key(df_a, key_cols_a, norm=key_norm),
                       val_a: pd.to_numeric(df_a[num_col_a], errors="coerce")})
    tb = pd.DataFrame({"KEY": _composite_key(df_b, key_cols_b, norm=key_norm),
                       val_b: pd.to_numeric(df_b[num_col_b], errors="coerce")})

    rows_a, rows_b = len(ta), len(tb)
    # Surface duplicate keys BEFORE collapsing them — an inner merge would
    # otherwise fan out (N×M), and keeping only the first row silently discards
    # the rest. Reviewers need to know which keys were affected.
    dups_a = _dup_keys(ta["KEY"], key_disp_a)
    dups_b = _dup_keys(tb["KEY"], key_disp_b)
    # Collapse duplicate keys per the chosen strategy: keep the first row, or
    # aggregate the value column (sum/mean/max/min) — sum is typical for money.
    if on_duplicate in ("sum", "mean", "max", "min"):
        ta = ta.groupby("KEY", as_index=False)[val_a].agg(on_duplicate)
        tb = tb.groupby("KEY", as_index=False)[val_b].agg(on_duplicate)
    else:
        ta = ta.drop_duplicates("KEY")
        tb = tb.drop_duplicates("KEY")

    merged             = pd.merge(ta, tb, on="KEY", how="inner")
    merged["Difference"] = merged[val_a] - merged[val_b]

    # The unmatched keys are the reconciliation "breaks": keys present on one
    # side but with no counterpart on the other. Surface them as their own
    # sheets — they're usually the whole point of running the comparison.
    a_keys, b_keys = set(ta["KEY"]), set(tb["KEY"])
    unmatched_a = ta[~ta["KEY"].isin(b_keys)].rename(columns={"KEY": key_disp_a})
    unmatched_b = tb[~tb["KEY"].isin(a_keys)].rename(columns={"KEY": key_disp_b})

    merged             = merged.rename(columns={"KEY": key_disp_a})

    a_gt_b    = int((merged["Difference"] > 0).sum())
    a_lt_b    = int((merged["Difference"] < 0).sum())
    zero_diff = int((merged["Difference"] == 0).sum())
    mean_diff = float(merged["Difference"].mean()) if not merged.empty else 0.0

    sh_unm_a = _sanitize_sheet(f"Unmatched {name_a}")
    sh_unm_b = _sanitize_sheet(f"Unmatched {name_b}")

    extra_frames: dict[str, pd.DataFrame] = {}
    if not dups_a.empty:
        extra_frames[_sanitize_sheet(f"Duplicate Keys {name_a}")] = dups_a
    if not dups_b.empty:
        extra_frames[_sanitize_sheet(f"Duplicate Keys {name_b}")] = dups_b

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        merged.to_excel(w, sheet_name="Differences", index=False)
        unmatched_a.to_excel(w, sheet_name=sh_unm_a, index=False)
        unmatched_b.to_excel(w, sheet_name=sh_unm_b, index=False)
        for sh, fr in extra_frames.items():
            fr.to_excel(w, sheet_name=sh, index=False)
    buf.seek(0)

    return {
        "type": "calc_diff",
        "metrics": {
            "rows_a": rows_a, "rows_b": rows_b,
            "matched": len(merged),
            "match_rate": round(len(merged) / max(len(ta), 1) * 100, 1),
            "a_gt_b": a_gt_b, "a_lt_b": a_lt_b, "zero": zero_diff,
            "mean_diff": round(mean_diff, 4),
            "excluded_a": len(ta) - len(merged),
            "excluded_b": len(tb) - len(merged),
            "unmatched_a": len(unmatched_a), "unmatched_b": len(unmatched_b),
            "duplicate_keys_a": len(dups_a), "duplicate_keys_b": len(dups_b),
        },
        "preview": {
            "differences": {"data": _to_records(merged), "total": len(merged),
                            "columns": merged.columns.tolist()},
            "unmatched_a": {"data": _to_records(unmatched_a), "total": len(unmatched_a),
                            "columns": unmatched_a.columns.tolist()},
            "unmatched_b": {"data": _to_records(unmatched_b), "total": len(unmatched_b),
                            "columns": unmatched_b.columns.tolist()},
        },
        "result_id": _save_result(buf.getvalue(), "CalcDifference_Result.xlsx"),
        "_frames": {"Differences": merged, sh_unm_a: unmatched_a, sh_unm_b: unmatched_b,
                    **extra_frames},
    }


# ── Custom Rule (no-code rule builder) ─────────────────────────────────────────
#
# config = {
#   axel_label, dms_label,
#   key_axel, key_dms,           # join key (key_dms optional → same as key_axel)
#   checks: [                    # one or more column checks, combined with AND
#     { axel_col, dms_col, mode: "numeric"|"text", op, tolerance }
#   ]
# }

_NUMERIC_OPS = {
    "eq":  lambda a, b, tol: (a - b).abs() <= tol,
    "ne":  lambda a, b, tol: (a - b).abs() >  tol,
    "gt":  lambda a, b, tol: a >  b + tol,
    "lt":  lambda a, b, tol: a <  b - tol,
    "gte": lambda a, b, tol: a >= b - tol,
    "lte": lambda a, b, tol: a <= b + tol,
}

_NUMERIC_LABEL = {
    "eq": "= (±tol)", "ne": "≠", "gt": ">", "lt": "<", "gte": "≥", "lte": "≤",
}

_TEXT_LABEL = {
    "eq": "=", "ne": "≠", "contains": "contains", "not_contains": "does not contain",
}


def _text_op(op: str, a: pd.Series, b: pd.Series) -> pd.Series:
    if op == "eq":
        return a == b
    if op == "ne":
        return a != b
    if op == "contains":
        return pd.Series([str(bb) in str(aa) for aa, bb in zip(a, b)], index=a.index)
    if op == "not_contains":
        return pd.Series([str(bb) not in str(aa) for aa, bb in zip(a, b)], index=a.index)
    raise HTTPException(400, f"Unknown text operator: {op}")


def run_custom_rule(df_a: pd.DataFrame, df_b: pd.DataFrame, config: dict) -> dict:
    name_a = config.get("axel_label", "AXEL")
    name_b = config.get("dms_label",  "DMS")
    # Join key: single column or composite (list). DMS mirrors AXEL when blank.
    key_cols_a = _key_cols(config.get("key_axel"))
    key_cols_b = _key_cols(config.get("key_dms")) or key_cols_a
    checks = config.get("checks", [])

    if not key_cols_a:
        raise HTTPException(400, "A join key column is required for the rule.")
    if not checks:
        raise HTTPException(400, "Add at least one comparison check to the rule.")

    key_a, key_b = _key_label(key_cols_a), _key_label(key_cols_b)
    key_norm = config.get("key_norm")

    # Distinct columns referenced by the checks (resolved against each frame)
    axel_cols = list(dict.fromkeys(c.get("axel_col", "") for c in checks if c.get("axel_col")))
    dms_cols  = list(dict.fromkeys(c.get("dms_col",  "") for c in checks if c.get("dms_col")))
    for c in checks:
        if not c.get("axel_col") or not c.get("dms_col"):
            raise HTTPException(400, "Every check needs both an AXEL and a DMS column.")

    left  = df_a[[_find_col(df_a, c) for c in axel_cols]].copy()
    left.columns  = [f"{c} [{name_a}]" for c in axel_cols]
    left.insert(0, "__KEY__", _composite_key(df_a, key_cols_a, norm=key_norm).values)
    right = df_b[[_find_col(df_b, c) for c in dms_cols]].copy()
    right.columns = [f"{c} [{name_b}]" for c in dms_cols]
    right.insert(0, "__KEY__", _composite_key(df_b, key_cols_b, norm=key_norm).values)

    rows_a, rows_b = len(left), len(right)
    # Surface duplicate keys before collapsing them (see calc_diff).
    dups_a = _dup_keys(left["__KEY__"], key_a)
    dups_b = _dup_keys(right["__KEY__"], key_b)
    # Keep first occurrence per key so duplicate keys don't fan out.
    left  = left.drop_duplicates("__KEY__")
    right = right.drop_duplicates("__KEY__")
    # Distinct-key counts drive the unmatched metrics: matched is the number of
    # distinct joined keys, so subtracting it from the raw row count (which may
    # include duplicate keys) would report phantom unmatched rows.
    distinct_a, distinct_b = len(left), len(right)

    # Unmatched keys = the reconciliation breaks (a key on one side with no
    # counterpart on the other). Carry the check columns so the sheet is useful.
    a_keys, b_keys = set(left["__KEY__"]), set(right["__KEY__"])
    unmatched_a = left[~left["__KEY__"].isin(b_keys)].rename(columns={"__KEY__": key_a})
    unmatched_b = right[~right["__KEY__"].isin(a_keys)].rename(columns={"__KEY__": key_b})

    merged = left.merge(right, on="__KEY__", how="inner")
    pass_all = pd.Series(True, index=merged.index)
    check_meta: list[tuple[str, str, pd.Series]] = []   # (col_a, col_b, fail_mask)

    for i, chk in enumerate(checks, start=1):
        col_a = f'{chk["axel_col"]} [{name_a}]'
        col_b = f'{chk["dms_col"]} [{name_b}]'
        op    = chk.get("op", "eq")
        mode  = chk.get("mode", "numeric")

        if mode == "numeric":
            if op not in _NUMERIC_OPS:
                raise HTTPException(400, f"Unknown numeric operator: {op}")
            tol = float(chk.get("tolerance") or 0)
            va  = pd.to_numeric(merged[col_a], errors="coerce")
            vb  = pd.to_numeric(merged[col_b], errors="coerce")
            res = _NUMERIC_OPS[op](va, vb, tol).fillna(False)
            merged[f"Δ{i} ({chk['axel_col']}−{chk['dms_col']})"] = va - vb
            label = _NUMERIC_LABEL.get(op, op)
        else:
            va  = merged[col_a].astype(str).str.strip()
            vb  = merged[col_b].astype(str).str.strip()
            res = _text_op(op, va, vb)
            label = _TEXT_LABEL.get(op, op)

        merged[f"Check {i}: {chk['axel_col']} {label} {chk['dms_col']}"] = \
            res.map({True: "Pass", False: "Fail"})
        check_meta.append((col_a, col_b, ~res))
        pass_all &= res

    merged["Result"] = pass_all.map({True: "Pass", False: "Fail"})

    # Readable pointer to the offending columns per failing row — survives into
    # the combined "run all" report (which can't carry cell colouring).
    if check_meta:
        def _fail_cols(pos: int) -> str:
            names: list[str] = []
            for col_a, col_b, fail_mask in check_meta:
                if bool(fail_mask.iloc[pos]):
                    names += [col_a, col_b]
            return ", ".join(names)
        merged["Failing Columns"] = [_fail_cols(p) for p in range(len(merged))]

    merged = merged.rename(columns={"__KEY__": key_a})

    matched = len(merged)
    passed  = int(pass_all.sum())
    failed  = matched - passed
    failures = merged[pass_all.values == False] if matched else merged  # noqa: E712

    sh_unm_a = _sanitize_sheet(f"Unmatched {name_a}")
    sh_unm_b = _sanitize_sheet(f"Unmatched {name_b}")

    extra_frames: dict[str, pd.DataFrame] = {}
    if not dups_a.empty:
        extra_frames[_sanitize_sheet(f"Duplicate Keys {name_a}")] = dups_a
    if not dups_b.empty:
        extra_frames[_sanitize_sheet(f"Duplicate Keys {name_b}")] = dups_b

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        merged.to_excel(w, sheet_name="Rule_Results", index=False)
        if failed:
            failures.to_excel(w, sheet_name="Failures", index=False)
        unmatched_a.to_excel(w, sheet_name=sh_unm_a, index=False)
        unmatched_b.to_excel(w, sheet_name=sh_unm_b, index=False)
        for sh, fr in extra_frames.items():
            fr.to_excel(w, sheet_name=sh, index=False)
    buf.seek(0)

    # Highlight the exact cells that broke each check, so a reviewer can jump
    # straight to the offending column+row.
    wb  = load_workbook(buf)
    ws  = wb["Rule_Results"]
    RED = PatternFill(start_color="FECACA", end_color="FECACA", fill_type="solid")
    hdr = {cell.value: i + 1 for i, cell in enumerate(ws[1])}
    result_idx = hdr.get("Result")

    for r in range(2, ws.max_row + 1):
        pos   = r - 2                                    # 0-based row in `merged`
        failed_here = False
        for col_a, col_b, fail_mask in check_meta:
            if bool(fail_mask.iloc[pos]):
                failed_here = True
                for cname in (col_a, col_b):
                    ci = hdr.get(cname)
                    if ci:
                        ws.cell(row=r, column=ci).fill = RED
        if failed_here and result_idx:
            ws.cell(row=r, column=result_idx).fill = RED

    ws.freeze_panes    = "A2"
    ws.auto_filter.ref = ws.dimensions
    out = io.BytesIO(); wb.save(out); out.seek(0)

    return {
        "type": "custom_rule",
        "metrics": {
            "rows_a": rows_a, "rows_b": rows_b,
            "matched": matched, "passed": passed, "failed": failed,
            "pass_rate": round(passed / max(matched, 1) * 100, 1),
            "checks": len(checks),
            "unmatched_a": distinct_a - matched, "unmatched_b": distinct_b - matched,
            "duplicate_keys_a": len(dups_a), "duplicate_keys_b": len(dups_b),
        },
        "preview": {
            "results": {"data": _to_records(merged), "total": matched,
                        "columns": merged.columns.tolist()},
            "unmatched_a": {"data": _to_records(unmatched_a), "total": len(unmatched_a),
                            "columns": unmatched_a.columns.tolist()},
            "unmatched_b": {"data": _to_records(unmatched_b), "total": len(unmatched_b),
                            "columns": unmatched_b.columns.tolist()},
        },
        "result_id": _save_result(out.getvalue(), "CustomRule_Result.xlsx"),
        "_frames": {"Rule_Results": merged, sh_unm_a: unmatched_a, sh_unm_b: unmatched_b,
                    **extra_frames},
    }


# ── Aggregate / group-by comparison ───────────────────────────────────────────
#
# config = {
#   axel_label, dms_label,
#   group_axel, group_dms,         # group-by key (single col or composite list)
#   metric: sum|count|mean|min|max|nunique,
#   value_axel, value_dms,         # value column (not needed for `count`)
#   tolerance, tolerance_pct,      # pass band: abs(diff) <= max(tol, pct% of DMS)
#   key_norm,                      # same canonicalisation options as join keys
# }

_AGG_METRICS = {"sum", "count", "mean", "min", "max", "nunique"}


def run_agg_compare(df_a: pd.DataFrame, df_b: pd.DataFrame, config: dict) -> dict:
    name_a = config.get("axel_label", "AXEL")
    name_b = config.get("dms_label",  "DMS")
    group_a = _key_cols(config.get("group_axel"))
    group_b = _key_cols(config.get("group_dms")) or group_a
    metric  = (config.get("metric") or "sum").strip().lower()
    val_a   = (config.get("value_axel") or "").strip()
    val_b   = (config.get("value_dms")  or "").strip()
    tol     = float(config.get("tolerance") or 0)
    tol_pct = float(config.get("tolerance_pct") or 0)
    key_norm = config.get("key_norm")

    if not group_a:
        raise HTTPException(400, "A group-by column is required for aggregate comparison.")
    if metric not in _AGG_METRICS:
        raise HTTPException(400, f"Unknown metric: {metric}")
    if metric != "count" and not (val_a and val_b):
        raise HTTPException(400, f"The '{metric}' metric needs a value column on both sides.")

    gdisp_a, gdisp_b = _key_label(group_a), _key_label(group_b)

    def _agg(df: pd.DataFrame, group_cols: list[str], valcol: str) -> pd.Series:
        key = _composite_key(df, group_cols, norm=key_norm)
        frame = pd.DataFrame({"GROUP": key.values})
        if metric == "count":
            return frame.groupby("GROUP").size()
        frame["V"] = pd.to_numeric(df[_find_col(df, valcol)], errors="coerce").values
        return frame.groupby("GROUP")["V"].agg(metric)

    ga = _agg(df_a, group_a, val_a)
    gb = _agg(df_b, group_b, val_b)

    label = f"{metric}({val_a or 'rows'})"
    m_col_a, m_col_b = f"{label} [{name_a}]", f"{label} [{name_b}]"

    merged = pd.DataFrame({m_col_a: ga}).join(pd.DataFrame({m_col_b: gb}), how="inner")
    merged["Difference"] = merged[m_col_a] - merged[m_col_b]
    allow = pd.concat(
        [pd.Series(tol, index=merged.index), merged[m_col_b].abs() * (tol_pct / 100.0)],
        axis=1,
    ).max(axis=1)
    passed_mask = merged["Difference"].abs() <= allow
    merged["Result"] = passed_mask.map({True: "Pass", False: "Fail"})
    merged = merged.reset_index().rename(columns={"GROUP": gdisp_a})

    a_only = ga.index.difference(gb.index)
    b_only = gb.index.difference(ga.index)
    unmatched_a = pd.DataFrame({gdisp_a: list(a_only), m_col_a: ga.loc[a_only].values})
    unmatched_b = pd.DataFrame({gdisp_b: list(b_only), m_col_b: gb.loc[b_only].values})

    matched = len(merged)
    passed  = int(passed_mask.sum())

    sh_unm_a = _sanitize_sheet(f"Unmatched {name_a}")
    sh_unm_b = _sanitize_sheet(f"Unmatched {name_b}")

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        merged.to_excel(w, sheet_name="Aggregate", index=False)
        unmatched_a.to_excel(w, sheet_name=sh_unm_a, index=False)
        unmatched_b.to_excel(w, sheet_name=sh_unm_b, index=False)
    buf.seek(0)

    # Red-fill the failing groups so a reviewer can spot them at a glance.
    wb = load_workbook(buf)
    ws = wb["Aggregate"]
    RED = PatternFill(start_color="FECACA", end_color="FECACA", fill_type="solid")
    hdr = {cell.value: i + 1 for i, cell in enumerate(ws[1])}
    res_idx = hdr.get("Result")
    for r in range(2, ws.max_row + 1):
        if res_idx and ws.cell(row=r, column=res_idx).value == "Fail":
            for c in range(1, ws.max_column + 1):
                ws.cell(row=r, column=c).fill = RED
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    out = io.BytesIO(); wb.save(out); out.seek(0)

    return {
        "type": "agg_compare",
        "metrics": {
            "groups_a": int(ga.size), "groups_b": int(gb.size),
            "matched": matched, "passed": passed, "failed": matched - passed,
            "pass_rate": round(passed / max(matched, 1) * 100, 1),
            "metric": metric,
            "unmatched_a": len(unmatched_a), "unmatched_b": len(unmatched_b),
        },
        "preview": {
            "results": {"data": _to_records(merged), "total": matched,
                        "columns": merged.columns.tolist()},
            "unmatched_a": {"data": _to_records(unmatched_a), "total": len(unmatched_a),
                            "columns": unmatched_a.columns.tolist()},
            "unmatched_b": {"data": _to_records(unmatched_b), "total": len(unmatched_b),
                            "columns": unmatched_b.columns.tolist()},
        },
        "result_id": _save_result(out.getvalue(), "AggregateCompare_Result.xlsx"),
        "_frames": {"Aggregate": merged, sh_unm_a: unmatched_a, sh_unm_b: unmatched_b},
    }


# ── Condition dispatcher ───────────────────────────────────────────────────────

def run_condition(df_axel: pd.DataFrame, df_dms: pd.DataFrame, ctype: str, config: dict) -> dict:
    # Row filtering applies to every comparison type — restrict both files
    # before the type-specific logic runs.
    filters = config.get("filters") or {}
    if filters:
        df_axel = _apply_filters(df_axel, filters.get("axel"), filters.get("row_range_axel"))
        df_dms  = _apply_filters(df_dms,  filters.get("dms"),  filters.get("row_range_dms"))

    if ctype == "sheet_diff":
        pairs    = config.get("col_pairs", [])
        if not pairs:
            raise HTTPException(400, "No column pairs defined for this condition.")
        axel_cols = [p["axel"] for p in pairs]
        dms_cols  = [p["dms"]  for p in pairs]
        return run_sheet_difference(df_axel, df_dms, axel_cols, dms_cols)

    if ctype == "stacked":
        return run_stacked_comparison(
            df_axel, df_dms,
            config.get("axel_label", "AXEL"),
            config.get("dms_label",  "DMS"),
            config.get("control_axel", ""),
            config.get("control_dms"),
        )

    if ctype == "calc_diff":
        return run_calc_difference(
            df_axel, df_dms,
            config.get("axel_label", "AXEL"),
            config.get("dms_label",  "DMS"),
            config.get("key_axel",  ""),
            config.get("val_axel",  ""),
            config.get("val_dms",   ""),
            config.get("key_dms"),
            config.get("key_norm"),
            config.get("on_duplicate", "first"),
        )

    if ctype == "custom_rule":
        return run_custom_rule(df_axel, df_dms, config)

    if ctype == "agg_compare":
        return run_agg_compare(df_axel, df_dms, config)

    raise HTTPException(400, f"Unknown condition type: {ctype}")


# ── Combined "run all" report ─────────────────────────────────────────────────

def run_all_conditions(
    conditions: list[dict],
    df_axel: pd.DataFrame,
    df_dms: pd.DataFrame,
) -> tuple[str, list[dict]]:
    """Run every enabled condition and produce a single combined Excel."""
    enabled = [c for c in conditions if c.get("enabled", True)]
    if not enabled:
        raise HTTPException(400, "No enabled conditions to run.")

    results: list[dict] = []
    errors:  list[dict] = []

    for cond in enabled:
        try:
            res = run_condition(df_axel, df_dms, cond["type"], cond.get("config", {}))
            results.append({"condition": cond, "result": res})
        except Exception as e:
            errors.append({"condition": cond, "error": str(e)})

    type_labels = {"sheet_diff": "Sheet Difference",
                   "stacked":    "Stacked Comparison",
                   "calc_diff":  "Calculation Difference",
                   "custom_rule": "Custom Rule",
                   "agg_compare": "Aggregate Comparison"}

    # Build combined Excel
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        # ── Summary sheet ───────────────────────────────────────────────────
        summary_rows = []
        for i, item in enumerate(results):
            cond = item["condition"]
            m    = item["result"]["metrics"]
            row  = {
                "#": i + 1,
                "Condition": cond["name"],
                "Type": type_labels.get(cond["type"], cond["type"]),
                "Status": "✓ OK",
            }
            if cond["type"] == "sheet_diff":
                row.update({"Matched": m["matched"], "Only in AXEL": m["only_in_a"],
                             "Only in DMS": m["only_in_b"], "Match Rate %": m["match_rate"]})
            elif cond["type"] == "stacked":
                row.update({"Paired": m["paired"], "AXEL only": m["a_only"],
                             "DMS only": m["b_only"], "Pair Rate %": m["pair_rate"]})
            elif cond["type"] == "calc_diff":
                row.update({"Matched": m["matched"], "A>B": m["a_gt_b"],
                             "A<B": m["a_lt_b"], "Mean Diff": m["mean_diff"],
                             "Match Rate %": m["match_rate"]})
            elif cond["type"] == "custom_rule":
                row.update({"Matched": m["matched"], "Passed": m["passed"],
                             "Failed": m["failed"], "Pass Rate %": m["pass_rate"]})
            elif cond["type"] == "agg_compare":
                row.update({"Matched": m["matched"], "Passed": m["passed"],
                             "Failed": m["failed"], "Pass Rate %": m["pass_rate"]})
            summary_rows.append(row)

        for item in errors:
            summary_rows.append({
                "#": "—",
                "Condition": item["condition"]["name"],
                "Type": type_labels.get(item["condition"]["type"], item["condition"]["type"]),
                "Status": f"✗ {item['error']}"
            })

        pd.DataFrame(summary_rows).to_excel(writer, sheet_name="Summary", index=False)

        # ── Per-condition sheets (full data via _frames) ────────────────────
        used_names: set[str] = {"Summary"}
        for i, item in enumerate(results):
            cond   = item["condition"]
            frames = item["result"].get("_frames", {})
            for key, df in frames.items():
                base = _sanitize_sheet(f"{i+1} {cond['name']} {key}", max_len=31)
                sh, n = base, 2
                while sh in used_names:                 # guarantee unique sheet names
                    suffix = f"~{n}"
                    sh = base[: 31 - len(suffix)] + suffix
                    n += 1
                used_names.add(sh)
                df.to_excel(writer, sheet_name=sh, index=False)

    buf.seek(0)

    # Style Summary sheet
    wb = load_workbook(buf)
    ws = wb["Summary"]
    header_fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    for col in ws.columns:
        max_w = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_w + 4, 40)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    out = io.BytesIO(); wb.save(out); out.seek(0)
    result_id = _save_result(out.getvalue(), "ValidationReport.xlsx")

    return result_id, [
        {"condition_id": r["condition"]["id"],
         "condition_name": r["condition"]["name"],
         "validation_name": r["condition"].get("validation_name", ""),
         "type": r["condition"]["type"],
         "metrics": r["result"]["metrics"],
         "result_id": r["result"]["result_id"]}
        for r in results
    ] + [{"condition_id": e["condition"]["id"],
          "condition_name": e["condition"]["name"],
          "validation_name": e["condition"].get("validation_name", ""),
          "type": e["condition"]["type"],
          "error": e["error"]} for e in errors]
